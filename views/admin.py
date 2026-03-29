# ============================================================
# FICHIER : views/admin.py (VERSION FINALE & COMPLÈTE)
# ============================================================
import json
import hashlib
import secrets
import re
import logging
import io
import os
import filetype
import uuid
from io import BytesIO
from urllib.parse import urlparse
from html import escape
from datetime import date, datetime, timedelta
from collections import defaultdict
from functools import wraps

from flask import (Blueprint, render_template, request, redirect, url_for,
                   flash, session, jsonify, send_file, current_app, abort, make_response)
from markupsafe import Markup
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
from sqlalchemy.exc import IntegrityError, SQLAlchemyError
from sqlalchemy import func, select, or_
from sqlalchemy.orm import joinedload

# --- IMPORTS POUR EXPORT PDF (ReportLab) ---
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm, mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Flowable, Image as ReportLabImage
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.graphics.shapes import Drawing, Rect, Line
from reportlab.graphics import renderPDF

# --- IMPORTS POUR EXPORT EXCEL (OpenPyXL) ---
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as OpenPyXLImage
from openpyxl.worksheet.datavalidation import DataValidation

# Imports Locaux
from extensions import limiter, cache
from db import db, Utilisateur, Parametre, Objet, Armoire, Categorie, Fournisseur, Kit, KitObjet, Budget, Depense, Echeance, Historique, Etablissement, Reservation, Suggestion, DocumentReglementaire, InventaireArchive, Salle
from utils import calculate_license_key, admin_required, login_required, log_action, get_etablissement_params, allowed_file, validate_email, validate_password_strength, validate_url
from fpdf import FPDF
from services.security_service import SecurityService
from services.document_service import DocumentService, DocumentServiceError

from PIL import Image, UnidentifiedImageError
try:
    import pillow_heif
    pillow_heif.register_heif_opener()
    HEIF_SUPPORTED = True
except ImportError:
    HEIF_SUPPORTED = False

# ============================================================
# CONFIGURATION
# ============================================================
admin_bp = Blueprint('admin', __name__, template_folder='../templates', url_prefix='/admin')

MAX_EXPORT_LIMIT = 3000
MAX_FILE_SIZE = 10 * 1024 * 1024 # 10 Mo
MAX_IMPORT_ROWS = 5000
MAX_JSON_SIZE = 5 * 1024 * 1024

MAX_ARMOIRES_PER_ETAB = 50
MAX_DESC_LENGTH = 500
MAX_IMAGE_PIXELS = 1920
JPEG_QUALITY = 80
UPLOAD_SUBDIR = 'armoires'

# ============================================================
# UTILITAIRES DE SÉCURITÉ
# ============================================================
def hash_user_id(user_id):
    return hashlib.sha256(str(user_id).encode()).hexdigest()[:8]

    return True, ""

def sanitize_for_excel(value):
    if isinstance(value, str) and value.startswith(('=', '+', '-', '@')): return f"'{value}"
    return value

def generer_code_unique():
    for _ in range(100):
        code = f"LABFLOW-{secrets.token_hex(3).upper()}"
        nested = db.session.begin_nested()
        try:
            test_etab = Etablissement(nom="__TEST__", code_invitation=code)
            db.session.add(test_etab)
            db.session.flush()
            nested.rollback()
            return code
        except IntegrityError:
            nested.rollback()
            continue
    raise RuntimeError("Impossible de générer un code unique")

def json_serial(obj):
    if isinstance(obj, (datetime, date)): return obj.isoformat()
    raise TypeError(f"Type {type(obj)} non sérialisable")

def sanitize_filename(name):
    safe = re.sub(r'[^\w\s-]', '', str(name))
    cleaned = secure_filename(safe)
    return cleaned[:100] if cleaned else "Export"

def sanitize_filename_report(text):
    if not text: return "Rapport"
    return re.sub(r'[^\w\s-]', '', str(text)).strip().replace(' ', '_')

def sanitize_for_excel_report(text):
    if not text: return ""
    text = str(text)
    if text.startswith(('=', '+', '-', '@')): text = "'" + text
    return text

def _handle_armoire_image(file_obj):
    """
    Helper dédié admin : Traite, sécurise et convertit (HEIC->JPG) l'image.
    """
    try:
        img = Image.open(file_obj)
        img.load() # Force le chargement pour valider l'intégrité
    except (UnidentifiedImageError, Exception):
        raise ValueError("Fichier image invalide ou corrompu.")

    # Validation Format (Liste blanche incluant HEIC/HEIF)
    valid_formats = ['JPEG', 'JPG', 'PNG', 'WEBP']
    if HEIF_SUPPORTED:
        valid_formats.extend(['HEIF', 'HEIC'])
    if img.format is None or img.format.upper() not in valid_formats:
        msg = "Format non supporté. Utilisez JPG, PNG ou HEIC (iPhone)." if HEIF_SUPPORTED else "Format non supporté. Utilisez JPG ou PNG."
        raise ValueError(msg)

    # Nettoyage (Sécurité) & Conversion RGB
    if img.mode in ("RGBA", "P"):
        img = img.convert("RGB")
    
    img_clean = img.copy()

    # Redimensionnement (Optimisation stockage)
    if img_clean.width > MAX_IMAGE_PIXELS or img_clean.height > MAX_IMAGE_PIXELS:
        img_clean.thumbnail((MAX_IMAGE_PIXELS, MAX_IMAGE_PIXELS), Image.Resampling.LANCZOS)

    # Génération nom sécurisé (UUID) + Extension forcée .jpg
    filename = f"ARM_{uuid.uuid4().hex}.jpg"
    
    # Chemins
    base_upload_dir = os.path.join(current_app.root_path, 'static', 'uploads', UPLOAD_SUBDIR)
    full_path = os.path.abspath(os.path.join(base_upload_dir, filename))
    
    # Sécurité Path Traversal
    if not full_path.startswith(os.path.abspath(base_upload_dir)):
        raise ValueError("Erreur de sécurité (Path Traversal).")

    os.makedirs(base_upload_dir, exist_ok=True)
    
    # Sauvegarde en JPEG (Universel pour le web)
    img_clean.save(full_path, "JPEG", quality=JPEG_QUALITY, optimize=True)

    return f"uploads/{UPLOAD_SUBDIR}/{filename}", full_path

def _rollback_file(path):
    """Supprime le fichier si l'enregistrement DB échoue"""
    if path and os.path.exists(path):
        try: os.remove(path)
        except OSError: pass
            
 
# ============================================================
# GÉNÉRATEURS PDF / EXCEL
# ============================================================

class LogoGraphique(Flowable):
    def __init__(self, width=40, height=40, logo_path=None):
        Flowable.__init__(self)
        self.width = width
        self.height = height
        self.logo_path = logo_path

    def draw(self):
        if self.logo_path and os.path.exists(self.logo_path):
            try:
                from reportlab.lib.utils import ImageReader
                img = ImageReader(self.logo_path)
                self.canv.drawImage(img, 0, 0, width=self.width, height=self.height, preserveAspectRatio=True, mask='auto')
                return
            except Exception:
                pass
        # Fallback : logo LabFlow vectoriel
        self.canv.setFillColor(colors.HexColor('#1F3B73'))
        self.canv.rect(0, 0, 8, 15, fill=1, stroke=0)
        self.canv.rect(12, 0, 8, 25, fill=1, stroke=0)
        self.canv.setFillColor(colors.HexColor('#4facfe'))
        self.canv.rect(24, 0, 8, 35, fill=1, stroke=0)
        self.canv.setStrokeColor(colors.HexColor('#FFD700'))
        self.canv.setLineWidth(2)
        self.canv.line(-5, 5, 35, 40)

class LogoEtablissement(Flowable):
    def __init__(self, logo_path, width=80, height=80):
        Flowable.__init__(self)
        self.logo_path = logo_path
        self.width = width
        self.height = height
    def draw(self):
        try:
            from reportlab.lib.utils import ImageReader
            img = ImageReader(self.logo_path)
            self.canv.drawImage(img, 0, 0, width=self.width, height=self.height, preserveAspectRatio=True, mask='auto')
        except Exception:
            pass

def get_logo_etablissement_path(etablissement_id):
    """Retourne le chemin absolu du logo établissement si disponible."""
    from utils import get_etablissement_params
    try:
        params = get_etablissement_params(etablissement_id)
        logo_url = params.get('logo_url')
        if logo_url:
            # logo_url est de type 'uploads/logo_1.png'
            logo_path = os.path.join(current_app.root_path, 'static', logo_url.lstrip('/static/'))
            if os.path.exists(logo_path):
                return logo_path
    except Exception:
        pass
    return None

def ajouter_logo_excel(ws):
    """Ajoute le logo LabFlow dans le fichier Excel si disponible."""
    logo_path = os.path.join(current_app.root_path, 'static', 'logo.png')
    if os.path.exists(logo_path):
        try:
            img = OpenPyXLImage(logo_path)
            img.width = 50
            img.height = 50
            ws.add_image(img, 'A1')
            return True
        except Exception:
            return False
    return False

def generer_budget_pdf_pro(data_export, metadata):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=1.5*cm, leftMargin=1.5*cm, topMargin=1.5*cm, bottomMargin=1.5*cm, title=f"Budget {metadata['etablissement']}")
    elements = []
    styles = getSampleStyleSheet()
    LABFLOW_BLUE = colors.HexColor('#1F3B73')
    style_titre = ParagraphStyle('Titre', parent=styles['Heading1'], fontSize=22, textColor=LABFLOW_BLUE, alignment=TA_CENTER)
    style_normal = ParagraphStyle('Normal', parent=styles['Normal'], fontSize=10)
    style_cell = ParagraphStyle('Cell', parent=styles['Normal'], fontSize=9)
    logo_path = metadata.get('logo_path')
    logo = LogoGraphique(width=60, height=60, logo_path=logo_path)
    titre_bloc = [
        Paragraph(metadata['etablissement'], style_titre),
        Paragraph(f"Rapport budgétaire", ParagraphStyle('SousTitre', parent=styles['Normal'], fontSize=12, textColor=colors.gray))
    ]
    header_table = Table([[logo, titre_bloc]], colWidths=[2*cm, 14*cm])
    header_table.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('LEFTPADDING', (0,0), (-1,-1), 0)]))
    elements.append(header_table)
    elements.append(Spacer(1, 0.5*cm))
    elements.append(Paragraph(f"Rapport du {metadata['date_debut']} au {metadata['date_fin']}", style_normal))
    elements.append(Spacer(1, 0.5*cm))
    table_data = [['Date', 'Fournisseur', 'Libellé', 'Montant']]
    for item in data_export:
        table_data.append([Paragraph(item['date'], style_cell), Paragraph(escape(item['fournisseur']), style_cell), Paragraph(escape(item['contenu']), style_cell), Paragraph(f"{item['montant']:.2f} €", style_cell)])
    table_data.append(['', '', 'TOTAL', f"{metadata['total']:.2f} €"])
    t = Table(table_data, colWidths=[2.5*cm, 5*cm, 8*cm, 3*cm])
    t.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), LABFLOW_BLUE), ('TEXTCOLOR', (0, 0), (-1, 0), colors.white), ('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), ('GRID', (0, 0), (-1, -1), 0.5, colors.grey), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.whitesmoke])]))
    elements.append(t)
    doc.build(elements)
    buffer.seek(0)
    filename = f"Budget_{sanitize_filename(metadata['etablissement'])}.pdf"
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/pdf')

def generer_budget_excel_pro(data_export, metadata):
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget"
    
    COLOR_PRIMARY = "1F3B73"
    fill_header = PatternFill(start_color=COLOR_PRIMARY, end_color=COLOR_PRIMARY, fill_type="solid")
    font_header = Font(name='Segoe UI', size=11, bold=True, color="FFFFFF")
    
    font_titre = Font(name='Segoe UI', size=16, bold=True, color=COLOR_PRIMARY)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")
    font_data = Font(name='Segoe UI', size=10)
    border_thin = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))

    ws.merge_cells('B1:E1')
    ws['B1'] = f"Rapport Budgétaire - {metadata['etablissement']}"
    ws['B1'].font = font_titre
    ws['B1'].alignment = align_left
    ws.row_dimensions[1].height = 40
    
    ajouter_logo_excel(ws)

    ws.merge_cells('A2:D2')
    ws['A2'] = f"Période : {metadata['date_debut']} au {metadata['date_fin']}"
    ws['A2'].alignment = align_left
    ws.merge_cells('A3:D3')
    ws['A3'] = f"Généré le {metadata['date_generation']} | {metadata['nombre_depenses']} écritures"
    ws['A3'].alignment = align_left
    ws.row_dimensions[3].height = 20

    headers = ['Date', 'Fournisseur', 'Libellé', 'Montant']
    ws.append([]) 
    ws.append(headers) 
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_center
        cell.border = border_thin
    
    ws.row_dimensions[5].height = 25

    current_row = 6
    for item in data_export:
        ws.cell(row=current_row, column=1, value=item['date']).alignment = align_center
        ws.cell(row=current_row, column=2, value=item['fournisseur'])
        ws.cell(row=current_row, column=3, value=item['contenu'])
        c4 = ws.cell(row=current_row, column=4, value=item['montant'])
        c4.number_format = '#,##0.00 €'
        c4.alignment = align_right
        
        for col in range(1, 5):
            ws.cell(row=current_row, column=col).border = border_thin
            ws.cell(row=current_row, column=col).font = font_data
        current_row += 1

    ws.cell(row=current_row, column=3, value="TOTAL").font = Font(bold=True)
    ws.cell(row=current_row, column=3).alignment = align_right
    c_total = ws.cell(row=current_row, column=4, value=metadata['total'])
    c_total.font = Font(bold=True, color=COLOR_PRIMARY, size=12)
    c_total.number_format = '#,##0.00 €'

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"Budget_{sanitize_filename(metadata['etablissement'])}.xlsx"
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

def generer_rapport_pdf(data, metadata):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=1.0*cm, leftMargin=1.0*cm, topMargin=1.0*cm, bottomMargin=1.0*cm, title=f"Rapport - {metadata['etablissement']}")
    elements = []
    styles = getSampleStyleSheet()
    logo_path = metadata.get('logo_path')
    logo = LogoGraphique(width=60, height=60, logo_path=logo_path)
    titre_style = ParagraphStyle('Titre', parent=styles['Heading1'], fontSize=22, textColor=colors.HexColor('#1F3B73'), alignment=TA_LEFT)
    sous_titre_style = ParagraphStyle('SousTitre', parent=styles['Normal'], fontSize=12, textColor=colors.gray, alignment=TA_LEFT)
    titre_bloc = [Paragraph(f"RAPPORT D'ACTIVITÉ", titre_style), Paragraph(f"{escape(metadata['etablissement'])}", sous_titre_style)]
    header_table = Table([[logo, titre_bloc]], colWidths=[1.5*cm, 20*cm])
    header_table.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('LEFTPADDING', (0,0), (-1,-1), 0)]))
    elements.append(header_table)
    elements.append(Spacer(1, 0.5*cm))
    meta_style = ParagraphStyle('Meta', parent=styles['Normal'], fontSize=10, textColor=colors.black)
    meta_label = ParagraphStyle('MetaLabel', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#1F3B73'), fontName='Helvetica-Bold')
    meta_data = [[Paragraph('Période :', meta_label), Paragraph(metadata['periode'], meta_style), Paragraph('Généré le :', meta_label), Paragraph(metadata['date_generation'], meta_style)], [Paragraph('Total :', meta_label), Paragraph(f"{metadata['total']} entrées", meta_style), '', '']]
    meta_table = Table(meta_data, colWidths=[2.5*cm, 8*cm, 2.5*cm, 8*cm])
    meta_table.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F8F9FA')), ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor('#CCCCCC')), ('PADDING', (0,0), (-1,-1), 6), ('VALIGN', (0,0), (-1,-1), 'MIDDLE')]))
    elements.append(meta_table)
    elements.append(Spacer(1, 0.8*cm))
    header_style = ParagraphStyle('Header', parent=styles['Normal'], fontSize=10, textColor=colors.white, fontName='Helvetica-Bold', alignment=TA_CENTER)
    cell_style = ParagraphStyle('Cell', parent=styles['Normal'], fontSize=9, textColor=colors.black)
    headers = [Paragraph('Date', header_style), Paragraph('Heure', header_style), Paragraph('Utilisateur', header_style), Paragraph('Action', header_style), Paragraph('Objet', header_style), Paragraph('Détails', header_style)]
    table_data = [headers]
    for row in data:
        table_data.append([Paragraph(row['date'], cell_style), Paragraph(row['heure'], cell_style), Paragraph(escape(row['utilisateur']), cell_style), Paragraph(escape(row['action']), cell_style), Paragraph(escape(row['objet']), cell_style), Paragraph(escape(row['details']) if row['details'] else '-', cell_style)])
    col_widths = [2.5*cm, 1.5*cm, 4.0*cm, 3.0*cm, 6.5*cm, 10.0*cm]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F3B73')), ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'), ('TOPPADDING', (0, 0), (-1, 0), 10), ('BOTTOMPADDING', (0, 0), (-1, 0), 10), ('VALIGN', (0, 1), (-1, -1), 'TOP'), ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F4F6F9')]), ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E0E0E0')), ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#1F3B73'))]))
    elements.append(t)
    doc.build(elements)
    buffer.seek(0)
    filename = f"Rapport_{sanitize_filename_report(metadata['etablissement'])}_{date.today().strftime('%Y%m%d')}.pdf"
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/pdf')

def generer_rapport_excel(data, metadata):
    wb = Workbook()
    ws = wb.active
    ws.title = "Activité"
    
    COLOR_PRIMARY = "1F3B73"
    fill_header = PatternFill(start_color="1F3B73", end_color="1F3B73", fill_type="solid")
    font_header = Font(name='Segoe UI', size=11, bold=True, color="FFFFFF")
    
    font_titre = Font(name='Segoe UI', size=18, bold=True, color=COLOR_PRIMARY)
    align_center = Alignment(horizontal="center", vertical="center")
    font_data = Font(name='Segoe UI', size=10)
    align_top = Alignment(vertical="top", wrap_text=True)
    border_thin = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))

    ws.merge_cells('B1:F1')
    ws['B1'] = f"RAPPORT D'ACTIVITÉ - {metadata['etablissement']}"
    ws['B1'].font = font_titre
    ws['B1'].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 45
    
    ajouter_logo_excel(ws)
    
    ws.merge_cells('A2:F4')
    ws['A2'] = f"Période : {metadata['periode']}\nGénéré le : {metadata['date_generation']}\nTotal : {metadata['total']} enregistrements"
    ws['A2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    headers = ["Date", "Heure", "Utilisateur", "Action", "Objet", "Détails"]
    ws.append([])
    ws.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col_num)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_center
    
    ws.row_dimensions[6].height = 30
    
    for row in data:
        ws.append([
            row['date'], row['heure'], 
            sanitize_for_excel_report(row['utilisateur']), 
            sanitize_for_excel_report(row['action']), 
            sanitize_for_excel_report(row['objet']), 
            sanitize_for_excel_report(row['details'])
        ])
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 50
    
    for row in ws.iter_rows(min_row=7, max_row=ws.max_row):
        for cell in row:
            cell.font = font_data
            cell.alignment = align_top
            cell.border = border_thin

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"Rapport_{sanitize_filename_report(metadata['etablissement'])}_{date.today().strftime('%Y%m%d')}.xlsx"
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    

def generer_inventaire_excel(data, metadata):
    """Génère un Excel propre (Style Premium sans image)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventaire"
    
    # --- 1. DÉFINITION DES STYLES ---
    COLOR_PRIMARY = "1F3B73"      # Bleu Nuit
    COLOR_HEADER_TEXT = "FFFFFF"  # Blanc
    COLOR_ZEBRA = "F4F6F9"        # Gris très clair
    COLOR_BORDER = "D0D0D0"       # Gris bordure
    COLOR_ALERT = "DC3545"        # Rouge Alerte

    # Polices
    font_titre = Font(name='Segoe UI', size=18, bold=True, color=COLOR_PRIMARY)
    # Sous-titre en gris et italique
    font_meta = Font(name='Segoe UI', size=11, italic=True, color="666666")
    
    font_header = Font(name='Segoe UI', size=11, bold=True, color=COLOR_HEADER_TEXT)
    font_data = Font(name='Segoe UI', size=10, color="333333")
    font_alert = Font(name='Segoe UI', size=10, bold=True, color=COLOR_ALERT)
    
    # Remplissages
    fill_header = PatternFill(start_color=COLOR_PRIMARY, end_color=COLOR_PRIMARY, fill_type="solid")
    fill_zebra = PatternFill(start_color=COLOR_ZEBRA, end_color=COLOR_ZEBRA, fill_type="solid")
    
    # Alignements
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Bordures
    border_thin = Border(
        left=Side(style='thin', color=COLOR_BORDER),
        right=Side(style='thin', color=COLOR_BORDER),
        top=Side(style='thin', color=COLOR_BORDER),
        bottom=Side(style='thin', color=COLOR_BORDER)
    )

    # --- 2. EN-TÊTE (TEXTE SEULEMENT) ---
    
    # Ligne 1 : TITRE PRINCIPAL
    ws.merge_cells('A1:F1')
    ws['A1'] = f"ÉTAT DE L'INVENTAIRE - {metadata['etablissement']}"
    ws['A1'].font = font_titre
    ws['A1'].alignment = align_left
    ws.row_dimensions[1].height = 35
    
    # Ligne 2 : SOUS-TITRE (Date | Total)
    ws.merge_cells('A2:F2')
    ws['A2'] = f"Généré le : {metadata['date_generation']} | Total références : {metadata['total']}"
    ws['A2'].font = font_meta
    ws['A2'].alignment = align_left
    ws.row_dimensions[2].height = 25
    
    # Ligne 3 : Espace vide pour aérer
    ws.row_dimensions[3].height = 15

    # --- 3. TABLEAU : HEADERS (Ligne 4) ---
    headers = ["Catégorie", "Désignation", "Quantité", "Seuil", "Emplacement", "Péremption"]
    
    header_row = 4
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_center
        cell.border = border_thin
    
    ws.row_dimensions[header_row].height = 30

    # --- 4. TABLEAU : DONNÉES ---
    start_row = 5
    
    for i, row in enumerate(data):
        current_row = start_row + i
        
        # Insertion des valeurs
        ws.cell(row=current_row, column=1, value=row['categorie'])
        ws.cell(row=current_row, column=2, value=row['nom'])
        ws.cell(row=current_row, column=3, value=row['quantite'])
        ws.cell(row=current_row, column=4, value=row['seuil'])
        ws.cell(row=current_row, column=5, value=row['armoire'])
        ws.cell(row=current_row, column=6, value=row['peremption'])
        
        # Application du style
        ws.row_dimensions[current_row].height = 25
        
        for col in range(1, 7):
            cell = ws.cell(row=current_row, column=col)
            cell.border = border_thin
            cell.font = font_data
            
            # Alignement : Centre pour les chiffres/dates, Gauche pour le texte
            if col in [3, 4, 6]: 
                cell.alignment = align_center
            else:
                cell.alignment = align_left
            
            # Zebra Striping
            if current_row % 2 == 0:
                cell.fill = fill_zebra
        
        # Alerte Stock Bas (Rouge sur la quantité)
        if row['quantite'] <= row['seuil']:
            ws.cell(row=current_row, column=3).font = font_alert

    # --- 5. FINITIONS ---
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 18
    
    # Filtres automatiques
    ws.auto_filter.ref = f"A{header_row}:F{start_row + len(data) - 1}"
    
    # Figer les volets sous les titres
    ws.freeze_panes = f"A{start_row}"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"Inventaire_{sanitize_filename_report(metadata['etablissement'])}_{date.today().strftime('%Y%m%d')}.xlsx"
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ============================================================
# HELPER DE SUPPRESSION ARMOIRE SÉCURISÉE (Version Gold)
# ============================================================
def _safe_delete_old_image(relative_path: str | None, context_info: str = "") -> None:
    """
    Supprime une ancienne image du disque APRÈS validation DB.
    Args:
        relative_path: Chemin relatif stocké en DB (ex: 'uploads/armoires/xyz.jpg')
        context_info: Chaîne pour tracer l'action dans les logs (User ID, Etab ID)
    """
    # Point 11 : Gestion robuste (None ou vide)
    if not relative_path or not relative_path.strip(): 
        return

    try:
        base_dir = os.path.join(current_app.root_path, 'static', 'uploads', UPLOAD_SUBDIR)
        full_path = os.path.abspath(os.path.join(current_app.root_path, 'static', relative_path))
        
        # Sécurité Path Traversal
        if not full_path.startswith(os.path.abspath(base_dir)):
            # Point 2 : Log enrichi
            current_app.logger.warning(f"SECURITY: Tentative suppression hors dossier ({context_info}) : {full_path}")
            return

        if os.path.exists(full_path):
            os.remove(full_path)
            current_app.logger.info(f"Nettoyage image : {relative_path} ({context_info})")

    # Point 3 : Exception spécifique
    except (OSError, IOError) as e:
        current_app.logger.error(f"Erreur suppression fichier {relative_path}: {e}")



# ============================================================
# ROUTE PRINCIPALE
# ============================================================
@admin_bp.route("/")
@admin_required
def admin():
    etablissement_id = session.get('etablissement_id')
    etablissement = db.session.get(Etablissement, etablissement_id)
    
    if not etablissement:
        admin_hash = hash_user_id(session.get('user_id'))
        current_app.logger.critical(f"Admin access attempt invalid etab by admin_{admin_hash}")
        flash("Erreur critique : Établissement introuvable.", "error")
        return redirect(url_for('auth.login'))

    if not etablissement.code_invitation:
        try:
            etablissement.code_invitation = generer_code_unique()
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            current_app.logger.error("Erreur génération code invitation", exc_info=True)

    params = get_etablissement_params(etablissement_id)
    
    licence_info = {
        'is_pro': params.get('licence_statut') == 'PRO',
        'instance_id': params.get('instance_id', 'N/A'),
        'statut': params.get('licence_statut', 'FREE')
    }

    # --- RECUPERATION STATS SÉCURITÉ ---
    try:
        sec_service = SecurityService()
        securite_stats = sec_service.get_dashboard_stats(etablissement_id)
    except Exception as e:
        current_app.logger.error(f"Erreur chargement stats sécurité admin: {str(e)}")
        securite_stats = None
    # -----------------------------------

    nb_admins = db.session.execute(db.select(db.func.count()).select_from(Utilisateur).filter_by(etablissement_id=etablissement_id, role="admin")).scalar()
    return render_template("admin.html", 
                           nb_admins=nb_admins,
                           licence=licence_info, 
                           etablissement=etablissement,
                           securite_stats=securite_stats)


# ============================================================
# THÈME & PERSONNALISATION
# ============================================================
# ================================================================
# PACKS ONBOARDING
# ================================================================
@admin_bp.route("/packs", methods=["GET"])
@admin_required
def packs_onboarding():
    """Page de sélection des packs de matériel."""
    from static.data.packs_onboarding import PACKS_ONBOARDING
    breadcrumbs = [
        {'text': 'Tableau de Bord', 'url': url_for('inventaire.index')},
        {'text': 'Administration', 'url': url_for('admin.admin')},
        {'text': 'Packs de matériel', 'url': None}
    ]
    return render_template("admin_packs.html", packs=PACKS_ONBOARDING, breadcrumbs=breadcrumbs)


@admin_bp.route("/packs/importer/<pack_id>", methods=["POST"])
@admin_required
def importer_pack(pack_id):
    """Importe un pack : crée armoires, catégories et objets."""
    from static.data.packs_onboarding import PACKS_ONBOARDING
    etablissement_id = session.get('etablissement_id')

    pack = next((p for p in PACKS_ONBOARDING if p["id"] == pack_id), None)
    if not pack:
        return jsonify({"success": False, "error": "Pack introuvable"}), 404

    try:
        stats = {"armoires": 0, "categories": 0, "objets": 0, "ignores": 0}

        # 1. Créer les armoires (si elles n'existent pas déjà)
        armoires_map = {}
        for arm_data in pack["armoires"]:
            existing = db.session.execute(
                db.select(Armoire).filter_by(
                    nom=arm_data["nom"],
                    etablissement_id=etablissement_id
                )
            ).scalar_one_or_none()
            if existing:
                armoires_map[arm_data["nom"]] = existing
            else:
                new_arm = Armoire(
                    nom=arm_data["nom"],
                    description=arm_data["description"],
                    etablissement_id=etablissement_id
                )
                db.session.add(new_arm)
                db.session.flush()
                armoires_map[arm_data["nom"]] = new_arm
                stats["armoires"] += 1

        # 2. Créer les catégories (si elles n'existent pas déjà)
        categories_map = {}
        for cat_nom in pack["categories"]:
            existing = db.session.execute(
                db.select(Categorie).filter_by(
                    nom=cat_nom,
                    etablissement_id=etablissement_id
                )
            ).scalar_one_or_none()
            if existing:
                categories_map[cat_nom] = existing
            else:
                new_cat = Categorie(
                    nom=cat_nom,
                    etablissement_id=etablissement_id
                )
                db.session.add(new_cat)
                db.session.flush()
                categories_map[cat_nom] = new_cat
                stats["categories"] += 1

        # 3. Créer les objets (si ils n'existent pas déjà)
        for obj_data in pack["objets"]:
            existing = db.session.execute(
                db.select(Objet).filter_by(
                    nom=obj_data["nom"],
                    etablissement_id=etablissement_id
                )
            ).scalar_one_or_none()
            if existing:
                stats["ignores"] += 1
                continue

            armoire = armoires_map.get(obj_data["armoire"])
            categorie = categories_map.get(obj_data["categorie"])

            new_obj = Objet(
                nom=obj_data["nom"],
                type_objet=obj_data["type_objet"],
                etablissement_id=etablissement_id,
                armoire_id=armoire.id if armoire else None,
                categorie_id=categorie.id if categorie else None,
                quantite_physique=obj_data["quantite_physique"],
                seuil=obj_data["seuil"],
                unite=obj_data.get("unite", "unité"),
                is_cmr=obj_data.get("is_cmr", False),
                image_url=obj_data.get("image_url"),
                fds_url=obj_data.get("fds_url"),
                en_commande=False,
                traite=False,
            )

            # Champs spécifiques aux produits chimiques
            if obj_data["type_objet"] == "produit":
                new_obj.capacite_initiale = obj_data.get("capacite_initiale")
                new_obj.niveau_actuel = obj_data.get("niveau_actuel")
                new_obj.seuil_pourcentage = obj_data.get("seuil_pourcentage", 50)

            db.session.add(new_obj)
            stats["objets"] += 1

        db.session.commit()
        log_action('import_pack', f"Pack '{pack['nom']}' importé : {stats}")
        cache.delete(f"armoires_{etablissement_id}")
        cache.delete(f"categories_{etablissement_id}")
        return jsonify({
            "success": True,
            "message": f"Pack importé avec succès ! {stats['armoires']} armoires, {stats['categories']} catégories et {stats['objets']} objets créés. {stats['ignores']} éléments ignorés (déjà existants).",
            "stats": stats
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Erreur import pack: {e}", exc_info=True)
        return jsonify({"success": False, "error": "Erreur technique lors de l'import"}), 500

@admin_bp.route("/personnalisation", methods=["GET"])
@admin_required
def personnalisation_page():
    breadcrumbs = [
        {'text': 'Tableau de Bord', 'url': url_for('inventaire.index')},
        {'text': 'Administration', 'url': url_for('admin.admin')},
        {'text': 'Personnalisation', 'url': None}
    ]
    params = get_etablissement_params(session.get('etablissement_id'))
    return render_template("admin_personnalisation.html", breadcrumbs=breadcrumbs, params=params)


@admin_bp.route("/theme", methods=["POST"])
@admin_required
def sauvegarder_theme():
    etablissement_id = session.get('etablissement_id')
    try:
        # Couleur principale
        couleur = request.form.get('couleur_principale', '').strip()
        if couleur and couleur.startswith('#') and len(couleur) in [4, 7]:
            for cle, valeur in [
                ('couleur_principale', couleur),
                ('couleur_secondaire', couleur)
            ]:
                param = db.session.execute(
                    db.select(Parametre).filter_by(
                        etablissement_id=etablissement_id, cle=cle
                    )
                ).scalar_one_or_none()
                if param:
                    param.valeur = valeur
                else:
                    db.session.add(Parametre(
                        etablissement_id=etablissement_id,
                        cle=cle, valeur=valeur
                    ))

        # Upload logo
        logo_file = request.files.get('logo_file')
        if logo_file and logo_file.filename:
            ext = logo_file.filename.rsplit('.', 1)[-1].lower()
            if ext in ['png', 'jpg', 'jpeg', 'svg', 'webp']:
                filename = f"logo_{etablissement_id}.{ext}"
                upload_path = os.path.join(
                    current_app.root_path, 'static', 'uploads', filename
                )
                os.makedirs(os.path.dirname(upload_path), exist_ok=True)
                logo_file.save(upload_path)
                logo_url = f"uploads/{filename}"

                param = db.session.execute(
                    db.select(Parametre).filter_by(
                        etablissement_id=etablissement_id, cle='logo_url'
                    )
                ).scalar_one_or_none()
                if param:
                    param.valeur = logo_url
                else:
                    db.session.add(Parametre(
                        etablissement_id=etablissement_id,
                        cle='logo_url', valeur=logo_url
                    ))

        db.session.commit()
        # Invalider le cache theme
        from extensions import cache
        from utils import get_etablissement_params
        cache.delete_memoized(get_etablissement_params, etablissement_id)
        flash("Thème mis à jour avec succès.", "success")

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Erreur sauvegarde thème: {e}")
        flash("Erreur lors de la sauvegarde.", "error")

    return redirect(url_for('admin.personnalisation_page'))

@admin_bp.route("/supprimer_logo", methods=["GET", "POST"])
@admin_required
def supprimer_logo():
    etablissement_id = session.get('etablissement_id')
    try:
        param = db.session.execute(
            db.select(Parametre).filter_by(
                etablissement_id=etablissement_id, cle='logo_url'
            )
        ).scalar_one_or_none()
        if param:
            # Supprimer le fichier
            logo_path = os.path.join(current_app.root_path, 'static', param.valeur)
            if os.path.exists(logo_path):
                os.remove(logo_path)
            db.session.delete(param)
            db.session.commit()
            # Invalider cache
            from extensions import cache
            from utils import get_etablissement_params
            cache.delete_memoized(get_etablissement_params, etablissement_id)
            flash("Logo supprimé avec succès.", "success")
        else:
            flash("Aucun logo à supprimer.", "warning")
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Erreur suppression logo: {e}")
        flash("Erreur lors de la suppression.", "error")
    return redirect(url_for('admin.admin'))

# ================================================================
# EXPORT INVENTAIRE
# ================================================================
@admin_bp.route("/export-inventaire", methods=["GET"])
@admin_required
def export_inventaire_page():
    from db import Armoire, Categorie
    etablissement_id = session.get('etablissement_id')
    armoires = db.session.execute(
        db.select(Armoire).filter_by(etablissement_id=etablissement_id).order_by(Armoire.nom)
    ).scalars().all()
    categories = db.session.execute(
        db.select(Categorie).filter_by(etablissement_id=etablissement_id).order_by(Categorie.nom)
    ).scalars().all()
    breadcrumbs = [
        {'text': 'Tableau de Bord', 'url': url_for('inventaire.index')},
        {'text': 'Administration', 'url': url_for('admin.admin')},
        {'text': 'Export inventaire', 'url': None}
    ]
    return render_template("admin_export_inventaire.html",
                           breadcrumbs=breadcrumbs,
                           all_armoires=armoires,
                           all_categories=categories)


# ================================================================
# RESET ÉTABLISSEMENT
# ================================================================
@admin_bp.route("/reset", methods=["GET"])
@admin_required
def reset_etablissement_page():
    breadcrumbs = [
        {'text': 'Tableau de Bord', 'url': url_for('inventaire.index')},
        {'text': 'Administration', 'url': url_for('admin.admin')},
        {'text': 'Reset établissement', 'url': None}
    ]
    etablissement = db.session.execute(
        db.select(Etablissement).filter_by(id=session.get('etablissement_id'))
    ).scalar_one_or_none()
    return render_template("admin_reset.html", breadcrumbs=breadcrumbs, etablissement=etablissement)


@admin_bp.route("/reset", methods=["POST"])
@admin_required
def reset_etablissement():
    if not request.is_json:
        return jsonify({'success': False, 'error': 'JSON requis'}), 415
    etablissement_id = session.get('etablissement_id')
    data = request.get_json()
    confirmation = data.get('confirmation', '')
    elements = data.get('elements', [])

    # Vérification confirmation
    etablissement = db.session.execute(
        db.select(Etablissement).filter_by(id=etablissement_id)
    ).scalar_one_or_none()
    if not etablissement or confirmation != etablissement.nom:
        return jsonify({'success': False, 'error': 'Confirmation incorrecte'}), 400

    try:
        from db import KitObjet, ReservationRecurrence, Salle
        stats = {}

        if 'reservations' in elements:
            count = db.session.execute(
                db.select(db.func.count(Reservation.id))
                .where(Reservation.etablissement_id == etablissement_id)
            ).scalar()
            db.session.execute(
                db.delete(Reservation).where(Reservation.etablissement_id == etablissement_id)
            )
            db.session.execute(
                db.delete(ReservationRecurrence).where(ReservationRecurrence.etablissement_id == etablissement_id)
            )
            stats['reservations'] = count

        if 'kits' in elements:
            kits = db.session.execute(
                db.select(Kit).filter_by(etablissement_id=etablissement_id)
            ).scalars().all()
            for kit in kits:
                db.session.execute(db.delete(KitObjet).where(KitObjet.kit_id == kit.id))
                db.session.delete(kit)
            stats['kits'] = len(kits)

        if 'inventaire' in elements:
            objets = db.session.execute(
                db.select(Objet).filter_by(etablissement_id=etablissement_id)
            ).scalars().all()
            for obj in objets:
                db.session.execute(db.delete(KitObjet).where(KitObjet.objet_id == obj.id))
                db.session.execute(db.delete(Reservation).where(Reservation.objet_id == obj.id))
                db.session.delete(obj)
            stats['inventaire'] = len(objets)

        if 'armoires' in elements:
            count = db.session.execute(
                db.select(db.func.count(Armoire.id))
                .where(Armoire.etablissement_id == etablissement_id)
            ).scalar()
            db.session.execute(
                db.delete(Armoire).where(Armoire.etablissement_id == etablissement_id)
            )
            stats['armoires'] = count

        if 'categories' in elements:
            count = db.session.execute(
                db.select(db.func.count(Categorie.id))
                .where(Categorie.etablissement_id == etablissement_id)
            ).scalar()
            db.session.execute(
                db.delete(Categorie).where(Categorie.etablissement_id == etablissement_id)
            )
            stats['categories'] = count

        if 'salles' in elements:
            count = db.session.execute(
                db.select(db.func.count(Salle.id))
                .where(Salle.etablissement_id == etablissement_id)
            ).scalar()
            db.session.execute(
                db.delete(Salle).where(Salle.etablissement_id == etablissement_id)
            )
            stats['salles'] = count

        if 'fournisseurs' in elements:
            count = db.session.execute(
                db.select(db.func.count(Fournisseur.id))
                .where(Fournisseur.etablissement_id == etablissement_id)
            ).scalar()
            db.session.execute(
                db.delete(Fournisseur).where(Fournisseur.etablissement_id == etablissement_id)
            )
            stats['fournisseurs'] = count

        if 'budget' in elements:
            budgets = db.session.execute(
                db.select(Budget).filter_by(etablissement_id=etablissement_id)
            ).scalars().all()
            for budget in budgets:
                db.session.execute(db.delete(Depense).where(Depense.budget_id == budget.id))
                db.session.delete(budget)
            stats['budget'] = len(budgets)

        db.session.commit()
        log_action('reset_etablissement', f"Reset: {list(elements)}")
        return jsonify({'success': True, 'stats': stats})

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Erreur reset: {e}", exc_info=True)
        return jsonify({'success': False, 'error': 'Erreur technique'}), 500

# ============================================================
# GESTION DES SALLES
# ============================================================
@admin_bp.route("/salles")
@admin_required
def gestion_salles():
    etablissement_id = session.get('etablissement_id')
    salles = db.session.execute(
        db.select(Salle)
        .filter_by(etablissement_id=etablissement_id)
        .order_by(Salle.nom)
    ).scalars().all()
    return render_template("admin_salles.html", salles=salles, breadcrumbs=[{'text': 'Tableau de Bord', 'url': url_for('inventaire.index')}, {'text': 'Administration', 'url': url_for('admin.admin')}, {'text': 'Gestion des salles', 'url': None}])

@admin_bp.route("/salles/ajouter", methods=["POST"])
@admin_required
def ajouter_salle():
    etablissement_id = session.get('etablissement_id')
    nom = request.form.get('nom', '').strip()
    description = request.form.get('description', '').strip()
    capacite = request.form.get('capacite', '').strip()

    if not nom:
        flash("Le nom de la salle est obligatoire.", "error")
        return redirect(url_for('admin.gestion_salles'))

    try:
        salle = Salle(
            nom=nom,
            description=description or None,
            capacite=int(capacite) if capacite.isdigit() else None,
            etablissement_id=etablissement_id
        )
        db.session.add(salle)
        db.session.commit()
        flash(f"Salle '{nom}' ajoutée.", "success")
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Erreur ajout salle: {e}")
        flash("Erreur lors de l'ajout.", "error")

    return redirect(url_for('admin.gestion_salles'))

@admin_bp.route("/salles/modifier/<int:salle_id>", methods=["POST"])
@admin_required
def modifier_salle(salle_id):
    etablissement_id = session.get('etablissement_id')
    salle = db.session.get(Salle, salle_id)
    if not salle or salle.etablissement_id != etablissement_id:
        flash("Salle introuvable.", "error")
        return redirect(url_for('admin.gestion_salles'))
    try:
        salle.nom = request.form.get('nom', '').strip() or salle.nom
        salle.description = request.form.get('description', '').strip() or None
        cap = request.form.get('capacite', '').strip()
        salle.capacite = int(cap) if cap.isdigit() else None
        db.session.commit()
        flash(f"Salle '{salle.nom}' modifiée.", "success")
    except Exception as e:
        db.session.rollback()
        flash("Erreur modification.", "error")
    return redirect(url_for('admin.gestion_salles'))

@admin_bp.route("/salles/supprimer/<int:salle_id>", methods=["POST"])
@admin_required
def supprimer_salle(salle_id):
    etablissement_id = session.get('etablissement_id')
    salle = db.session.get(Salle, salle_id)

    if not salle or salle.etablissement_id != etablissement_id:
        flash("Salle introuvable.", "error")
        return redirect(url_for('admin.gestion_salles'))

    try:
        # Détacher les réservations liées
        db.session.execute(
            db.update(Reservation)
            .where(Reservation.salle_id == salle_id)
            .values(salle_id=None)
        )
        db.session.delete(salle)
        db.session.commit()
        flash(f"Salle '{salle.nom}' supprimée.", "success")
    except Exception as e:
        db.session.rollback()
        flash("Erreur suppression.", "error")

    return redirect(url_for('admin.gestion_salles'))

# ============================================================
# GESTION CREATION CATÉGORIES
# ============================================================
@admin_bp.route("/categories/ajouter", methods=["POST"])
@admin_required
@limiter.limit("30 per minute")
def ajouter_categorie():
    etablissement_id = session['etablissement_id']
    nom = request.form.get("nom", "").strip()
    
    if not nom:
        flash("Le nom ne peut pas être vide.", "error")
        return redirect(url_for('main.gestion_categories'))
    
    try:
        nouvelle_categorie = Categorie(nom=nom, etablissement_id=etablissement_id)
        db.session.add(nouvelle_categorie)
        db.session.commit()
        flash(f"La catégorie '{nom}' a été créée.", "success")
    except IntegrityError:
        db.session.rollback()
        flash(f"La catégorie '{nom}' existe déjà.", "error")
    except Exception:
        db.session.rollback()
        current_app.logger.error(f"Erreur ajout catégorie", exc_info=True)
        flash("Erreur technique.", "error")
    
    return redirect(url_for('main.gestion_categories'))
    
# ===================================================================
# Ajout spécifique pour les armoires avec photo et description
# ===================================================================
@admin_bp.route("/armoires/ajouter_specifique", methods=["POST"])
@admin_required
@limiter.limit("10 per minute")
def ajouter_armoire_specifique():
    etablissement_id = session['etablissement_id']
    user_id = session.get('user_id', 'Unknown')
    
    # 1. Validation Quotas
    count = db.session.query(func.count(Armoire.id)).filter_by(etablissement_id=etablissement_id).scalar()
    if count >= MAX_ARMOIRES_PER_ETAB:
        flash(f"Limite atteinte ({MAX_ARMOIRES_PER_ETAB} armoires max).", "warning")
        return redirect(url_for('main.gestion_armoires'))

    # 2. Validation Formulaire
    nom = request.form.get("nom", "").strip()
    description = request.form.get("description", "").strip() or None

    if not nom:
        flash("Le nom est obligatoire.", "error")
        return redirect(url_for('main.gestion_armoires'))
    
    if len(nom) > 100:
        flash("Nom trop long.", "warning")
        return redirect(url_for('main.gestion_armoires'))

    if description and len(description) > MAX_DESC_LENGTH:
        flash("Description trop longue.", "warning")
        return redirect(url_for('main.gestion_armoires'))

    # 3. Gestion Image (HEIC -> JPG)
    photo_db_path = None
    uploaded_file_path = None

    if 'photo' in request.files:
        file = request.files['photo']
        if file and file.filename != '':
            try:
                photo_db_path, uploaded_file_path = _handle_armoire_image(file)
            except ValueError as ve:
                flash(str(ve), "warning")
                return redirect(url_for('main.gestion_armoires'))
            except Exception as e:
                current_app.logger.error(f"Upload error: {str(e)}", exc_info=True)
                flash("Erreur technique image.", "error")
                return redirect(url_for('main.gestion_armoires'))

    # 4. Enregistrement DB
    try:
        nouvelle_armoire = Armoire(
            nom=nom,
            description=description,
            photo_url=photo_db_path,
            etablissement_id=etablissement_id
        )
        db.session.add(nouvelle_armoire)
        db.session.commit()
        
        current_app.logger.info(f"Armoire créée : {nom} (ID: {nouvelle_armoire.id}) par {user_id}")
        flash(f"L'armoire '{nom}' a été créée.", "success")

    except IntegrityError:
        db.session.rollback()
        _rollback_file(uploaded_file_path)
        flash(f"Une armoire nommée '{nom}' existe déjà.", "warning")
    except Exception as e:
        db.session.rollback()
        _rollback_file(uploaded_file_path)
        current_app.logger.error(f"DB Error: {str(e)}", exc_info=True)
        flash("Erreur technique base de données.", "error")

    return redirect(url_for('main.gestion_armoires'))

@admin_bp.route("/supprimer/<type_objet>/<int:id>", methods=["POST"])
@admin_required
@limiter.limit("20 per minute")
def supprimer(type_objet, id):
    etablissement_id = session['etablissement_id']
    
    if type_objet == "armoire":
        Model = Armoire
        redirect_to = "main.gestion_armoires"
    elif type_objet == "categorie":
        Model = Categorie
        redirect_to = "main.gestion_categories"
    else:
        abort(400)

    element = db.session.get(Model, id)

    if not element or element.etablissement_id != etablissement_id:
        admin_hash = hash_user_id(session.get('user_id'))
        current_app.logger.warning(f"IDOR SUSPECT: AdminHash {admin_hash} delete {type_objet} {id}")
        abort(403)

    if type_objet == "armoire":
        count = db.session.query(Objet).filter_by(armoire_id=id, etablissement_id=etablissement_id).count()
    else:
        count = db.session.query(Objet).filter_by(categorie_id=id, etablissement_id=etablissement_id).count()
        
    if count > 0:
        flash(f"Impossible de supprimer '{element.nom}' car il contient encore {count} objet(s).", "error")
        return redirect(url_for(redirect_to))

    try:
        nom_element = element.nom
        db.session.delete(element)
        db.session.commit()
        flash(f"L'élément '{nom_element}' a été supprimé.", "success")
    except Exception:
        db.session.rollback()
        current_app.logger.error(f"Erreur suppression {type_objet}", exc_info=True)
        flash("Erreur technique.", "error")
    
    return redirect(url_for(redirect_to))

def _modifier_element_generique(model_class, request_data):
    """Helper pour modifier Armoire ou Categorie sans duplication."""
    try:
        etablissement_id = session['etablissement_id']
        element_id = request_data.get("id")
        nouveau_nom = request_data.get("nom", "").strip()

        if not all([element_id, nouveau_nom]): 
            return jsonify(success=False, error="Données invalides"), 400

        element = db.session.get(model_class, element_id)
        if not element or element.etablissement_id != etablissement_id:
            return jsonify(success=False, error="Élément introuvable"), 404

        element.nom = nouveau_nom
        db.session.commit()
        return jsonify(success=True, nouveau_nom=nouveau_nom)
    
    except IntegrityError:
        db.session.rollback()
        return jsonify(success=False, error="Ce nom existe déjà"), 409
    except Exception:
        db.session.rollback()
        current_app.logger.error(f"Erreur modif {model_class.__name__}", exc_info=True)
        return jsonify(success=False, error="Erreur serveur"), 500


# ============================================================
# ROUTE DE MODIFICATION ARMOIRES
# ============================================================
@admin_bp.route("/armoires/modifier_specifique", methods=["POST"])
@admin_required
def modifier_armoire_specifique():
    etablissement_id = session['etablissement_id']
    user_id = session.get('user_id', 'Unknown')
    context_log = f"User {user_id} | Etab {etablissement_id}"
    
    armoire_id = request.form.get("id")
    
    # Point 12 : Gestion Concurrence (Lock Pessimiste)
    # On verrouille la ligne jusqu'à la fin de la transaction
    try:
        armoire = db.session.query(Armoire).filter_by(id=armoire_id).with_for_update().first()
    except Exception as e:
        # Si la DB ne supporte pas le lock ou timeout
        current_app.logger.error(f"DB Lock Error: {e}")
        flash("Erreur technique (verrouillage). Réessayez.", "error")
        return redirect(url_for('main.gestion_armoires'))

    # Point 8 : IDOR Message standardisé
    if not armoire or armoire.etablissement_id != etablissement_id:
        flash("Armoire introuvable.", "error")
        return redirect(url_for('main.gestion_armoires'))

    # --- Validation ---
    nom = request.form.get("nom", "").strip()
    description = request.form.get("description", "").strip() or None
    # Point 1 : La logique checkbox est gérée, le JS gérera l'UX
    supprimer_photo = request.form.get("supprimer_photo") == "1"

    if not nom:
        flash("Le nom est obligatoire.", "error")
        return redirect(url_for('main.gestion_armoires'))
    
    if len(nom) > 100:
        flash("Nom trop long.", "warning")
        return redirect(url_for('main.gestion_armoires'))

    if description and len(description) > MAX_DESC_LENGTH:
        flash("Description trop longue.", "warning")
        return redirect(url_for('main.gestion_armoires'))

    # --- Point 5 : Détection "Pas de changement" ---
    # On vérifie si l'utilisateur a cliqué sur "Enregistrer" sans rien toucher
    no_text_change = (nom == armoire.nom and description == armoire.description)
    no_photo_action = (not supprimer_photo and 'photo' not in request.files)
    # Cas subtil : upload vide
    if 'photo' in request.files and request.files['photo'].filename == '':
        no_photo_action = True

    if no_text_change and no_photo_action:
        flash("Aucune modification détectée.", "info")
        return redirect(url_for('main.gestion_armoires'))

    # --- Vérification Unicité ---
    existing = db.session.query(Armoire.id).filter(
        func.lower(Armoire.nom) == nom.lower(),
        Armoire.etablissement_id == etablissement_id,
        Armoire.id != armoire.id 
    ).first()

    if existing:
        # Point 4 : Log du conflit
        current_app.logger.info(f"Conflit nom armoire '{nom}' avec ID {existing.id} ({context_log})")
        flash(f"Le nom '{nom}' est déjà utilisé.", "warning")
        return redirect(url_for('main.gestion_armoires'))

    # --- Logique Image ---
    old_photo_path = armoire.photo_url
    new_photo_path = old_photo_path
    uploaded_file_physical_path = None
    file_to_delete_later = None

    # Cas A : Nouvel Upload
    if 'photo' in request.files and request.files['photo'].filename != '':
        try:
            # Point 7 : Nom uniformisé (_handle_armoire_image)
            rel_path, abs_path = _handle_armoire_image(request.files['photo'])
            
            new_photo_path = rel_path
            uploaded_file_physical_path = abs_path
            
            if old_photo_path:
                file_to_delete_later = old_photo_path

        except ValueError as ve:
            flash(str(ve), "warning")
            return redirect(url_for('main.gestion_armoires'))
        except Exception as e:
            current_app.logger.error(f"Upload error: {e}", exc_info=True)
            flash("Erreur technique image.", "error")
            return redirect(url_for('main.gestion_armoires'))

    # Cas B : Suppression explicite
    elif supprimer_photo:
        new_photo_path = None
        if old_photo_path:
            file_to_delete_later = old_photo_path

    # --- Mise à jour ---
    armoire.nom = nom
    armoire.description = description
    armoire.photo_url = new_photo_path

    # --- Commit & Nettoyage ---
    try:
        db.session.commit()
        
        # SUCCÈS : Suppression différée
        if file_to_delete_later:
            _safe_delete_old_image(file_to_delete_later, context_log)
            
        current_app.logger.info(f"Armoire modifiée : ID {armoire.id} ({context_log})")
        # Point 8 : Message concis
        flash(f"L'armoire '{nom}' a été modifiée.", "success")

    # Point 9 : Distinction des erreurs
    except IntegrityError:
        db.session.rollback()
        if uploaded_file_physical_path: _rollback_file(uploaded_file_physical_path)
        flash("Erreur d'intégrité (doublon probable).", "error")

    except Exception as e:
        db.session.rollback()
        if uploaded_file_physical_path: _rollback_file(uploaded_file_physical_path)
        current_app.logger.error(f"DB Error update armoire: {e}", exc_info=True)
        flash("Erreur technique lors de l'enregistrement.", "error")

    return redirect(url_for('main.gestion_armoires'))

# ============================================================
# ROUTE DE MODIFICATION CATEGORIE
# ============================================================
@admin_bp.route("/modifier_categorie", methods=["POST"])
@admin_required
def modifier_categorie():
    return _modifier_element_generique(Categorie, request.get_json())