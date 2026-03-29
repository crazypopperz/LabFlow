# ============================================================
# FICHIER : views/admin_import.py
# Sauvegarde, restauration, import/export, rapports
# ============================================================
from flask import Blueprint, render_template, request, redirect, url_for, flash, session, current_app, send_file, jsonify
from werkzeug.utils import secure_filename as sanitize_filename
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import joinedload
from datetime import datetime, date
from io import BytesIO
import os
import json
import io

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from extensions import limiter, cache
from db import db, Objet, Armoire, Categorie, Utilisateur, Historique, Etablissement, Reservation, InventaireArchive, Fournisseur, Budget, Echeance, KitObjet, Suggestion, Depense, Kit
from utils import admin_required, log_action, allowed_file, get_etablissement_params, invalidate_alertes_cache

admin_import_bp = Blueprint('admin_import', __name__, url_prefix='/admin')

@admin_import_bp.route("/sauvegardes")
@admin_required
def gestion_sauvegardes():
    etablissement_id = session['etablissement_id']
    params = get_etablissement_params(etablissement_id)
    if params.get('licence_statut') != 'PRO':
        flash("Réservé à la version PRO.", "warning")
        return redirect(url_for('admin.admin'))
    return render_template("admin_backup.html", breadcrumbs=[{'text': 'Tableau de Bord', 'url': url_for('inventaire.index')}, {'text': 'Administration', 'url': url_for('admin.admin')}, {'text': 'Sauvegardes & Restauration', 'url': None}])

def json_serial(obj):
    import datetime
    if isinstance(obj, (datetime.date, datetime.datetime)):
        return obj.isoformat()
    raise TypeError(f"Type {type(obj)} non sérialisable")

@admin_import_bp.route("/telecharger_db")
@admin_required
def telecharger_db():
    etablissement_id = session['etablissement_id']
    params = get_etablissement_params(etablissement_id)
    if params.get('licence_statut') != 'PRO':
        flash("Réservé PRO.", "warning")
        return redirect(url_for('admin_import.gestion_sauvegardes'))

    try:
        data = {
            'metadata': {'version': '1.0', 'date': datetime.now().isoformat(), 'etablissement': session.get('nom_etablissement')},
            'armoires': [], 'categories': [], 'fournisseurs': [], 'objets': [], 'budget': [], 'depenses': [], 'echeances': []
        }
        
        # Extraction des données (simplifiée pour la lisibilité)
        for a in db.session.execute(db.select(Armoire).filter_by(etablissement_id=etablissement_id)).scalars():
            data['armoires'].append({'id': a.id, 'nom': a.nom})
        for c in db.session.execute(db.select(Categorie).filter_by(etablissement_id=etablissement_id)).scalars():
            data['categories'].append({'id': c.id, 'nom': c.nom})
        for f in db.session.execute(db.select(Fournisseur).filter_by(etablissement_id=etablissement_id)).scalars():
            data['fournisseurs'].append({'id': f.id, 'nom': f.nom, 'site_web': f.site_web, 'logo': f.logo})
        for o in db.session.execute(db.select(Objet).filter_by(etablissement_id=etablissement_id)).scalars():
            data['objets'].append({
                'id': o.id, 'nom': o.nom, 'type_objet': o.type_objet, 'quantite': o.quantite_physique, 'seuil': o.seuil,
                'armoire_id': o.armoire_id, 'categorie_id': o.categorie_id, 'date_peremption': o.date_peremption,
                'image_url': o.image_url, 'fds_url': o.fds_url
            })
        for b in db.session.execute(db.select(Budget).filter_by(etablissement_id=etablissement_id)).scalars():
            b_data = {'id': b.id, 'annee': b.annee, 'montant': b.montant_initial, 'cloture': b.cloture}
            data['budget'].append(b_data)
            for d in b.depenses:
                data['depenses'].append({
                    'budget_id': b.id, 'date': d.date_depense, 'contenu': d.contenu,
                    'montant': d.montant, 'est_bon_achat': d.est_bon_achat, 'fournisseur_id': d.fournisseur_id
                })
        for e in db.session.execute(db.select(Echeance).filter_by(etablissement_id=etablissement_id)).scalars():
            data['echeances'].append({'intitule': e.intitule, 'date': e.date_echeance, 'details': e.details, 'traite': e.traite})

        json_str = json.dumps(data, default=json_serial, indent=4)
        buffer = BytesIO()
        buffer.write(json_str.encode('utf-8'))
        buffer.seek(0)
        log_action('backup_download', "Export JSON")
        safe_etab = sanitize_filename(session.get('nom_etablissement', 'Backup'))
        return send_file(buffer, as_attachment=True, download_name=f"Sauvegarde_{safe_etab}_{date.today()}.json", mimetype='application/json')

    except Exception:
        current_app.logger.error("Erreur backup", exc_info=True)
        flash("Erreur technique.", "error")
        return redirect(url_for('admin_import.gestion_sauvegardes'))

MAX_JSON_SIZE = 10 * 1024 * 1024  # 10 Mo
@admin_import_bp.route("/importer_db", methods=["POST"])
@admin_required
def importer_db():
    etablissement_id = session['etablissement_id']
    if 'fichier' not in request.files:
        flash("Aucun fichier.", "error")
        return redirect(url_for('admin_import.gestion_sauvegardes'))
        
    fichier = request.files['fichier']
    
    # CORRECTION : Vérification taille fichier avant lecture
    fichier.seek(0, 2)
    if fichier.tell() > MAX_JSON_SIZE:
        flash("Fichier JSON trop volumineux (Max 5Mo).", "error")
        return redirect(url_for('admin_import.gestion_sauvegardes'))
    fichier.seek(0)

    if fichier.filename == '' or not allowed_file(fichier.filename) or not fichier.filename.endswith('.json'):
        flash("Fichier invalide (.json requis).", "error")
        return redirect(url_for('admin_import.gestion_sauvegardes'))

    try:
        data = json.load(fichier)
        
        # CORRECTION : Transaction atomique (Tout ou rien)
        with db.session.begin_nested():
            map_armoires = {}
            map_categories = {}
            map_fournisseurs = {}
            map_budgets = {}

            # Nettoyage
            for model in [KitObjet, Historique, Reservation, Suggestion, Depense, Objet, Kit, Budget, Echeance, Armoire, Categorie, Fournisseur]:
                db.session.query(model).filter_by(etablissement_id=etablissement_id).delete()
            db.session.flush()

            # Reconstruction
            for a in data.get('armoires', []):
                new_a = Armoire(nom=a['nom'], etablissement_id=etablissement_id)
                db.session.add(new_a)
                db.session.flush()
                map_armoires[a['id']] = new_a.id

            for c in data.get('categories', []):
                new_c = Categorie(nom=c['nom'], etablissement_id=etablissement_id)
                db.session.add(new_c)
                db.session.flush()
                map_categories[c['id']] = new_c.id

            for f in data.get('fournisseurs', []):
                new_f = Fournisseur(nom=f['nom'], site_web=f.get('site_web'), logo=f.get('logo'), etablissement_id=etablissement_id)
                db.session.add(new_f)
                db.session.flush()
                map_fournisseurs[f['id']] = new_f.id

            for o in data.get('objets', []):
                new_armoire_id = map_armoires.get(o['armoire_id'])
                new_cat_id = map_categories.get(o['categorie_id'])
                if not new_armoire_id or not new_cat_id: continue 
                
                date_perim = None
                if o.get('date_peremption'):
                    try: date_perim = datetime.fromisoformat(o['date_peremption']).date()
                    except: pass

                db.session.add(Objet(
                    nom=o['nom'], type_objet=o.get('type_objet', 'materiel'), quantite_physique=o['quantite'], seuil=o['seuil'],
                    armoire_id=new_armoire_id, categorie_id=new_cat_id, date_peremption=date_perim,
                    image_url=o.get('image_url'), fds_url=o.get('fds_url'), etablissement_id=etablissement_id
                ))

            for b in data.get('budget', []):
                new_b = Budget(annee=b['annee'], montant_initial=b['montant'], cloture=b['cloture'], etablissement_id=etablissement_id)
                db.session.add(new_b)
                db.session.flush()
                map_budgets[b['id']] = new_b.id

            for d in data.get('depenses', []):
                new_budget_id = map_budgets.get(d['budget_id'])
                if not new_budget_id: continue
                new_fournisseur_id = map_fournisseurs.get(d['fournisseur_id']) if d.get('fournisseur_id') else None
                try: date_dep = datetime.fromisoformat(d['date']).date()
                except: date_dep = date.today()
                db.session.add(Depense(
                    budget_id=new_budget_id, fournisseur_id=new_fournisseur_id, contenu=d['contenu'],
                    montant=d['montant'], date_depense=date_dep, est_bon_achat=d.get('est_bon_achat', False),
                    etablissement_id=etablissement_id
                ))

            for e in data.get('echeances', []):
                try: date_ech = datetime.fromisoformat(e['date']).date()
                except: continue
                db.session.add(Echeance(intitule=e['intitule'], date_echeance=date_ech, details=e.get('details'), traite=e.get('traite', 0), etablissement_id=etablissement_id))

        db.session.commit()
        log_action('backup_restore', "Import JSON")
        flash("Restauration réussie !", "success")

    except Exception:
        db.session.rollback()
        current_app.logger.error("Erreur import DB", exc_info=True)
        flash("Erreur critique lors de l'importation.", "error")

    return redirect(url_for('admin_import.gestion_sauvegardes'))

# ============================================================
# IMPORT EXCEL — ASSISTANT 3 ETAPES
# ============================================================

# Mots-clés pour la détection automatique des colonnes
MAPPING_KEYWORDS = {
    'nom': ['nom', 'designation', 'désignation', 'produit', 'article', 'libelle', 'libellé', 'intitule', 'intitulé', 'materiel', 'matériel'],
    'quantite': ['qte', 'quantite', 'quantité', 'stock', 'nombre', 'qté'],
    'seuil': ['seuil', 'alerte', 'minimum', 'mini', 'min'],
    'armoire': ['armoire', 'localisation', 'emplacement', 'lieu', 'salle', 'rangement'],
    'categorie': ['categorie', 'catégorie', 'famille', 'type', 'classe'],
    'date_peremption': ['peremption', 'péremption', 'expiration', 'dlc', 'dlu', 'date'],
    'type_objet': ['type_objet', 'type objet', 'nature'],
    'is_cmr': ['cmr', 'cancero', 'cancéro', 'dangereux'],
    'unite': ['unite', 'unité', 'unit'],
}

MAX_FILE_SIZE = 5 * 1024 * 1024  # 5 Mo
MAX_IMPORT_ROWS = 500


def detect_mapping(headers):
    """Détecte automatiquement le mapping colonnes -> champs LabFlow"""
    mapping = {}
    for field, keywords in MAPPING_KEYWORDS.items():
        for i, header in enumerate(headers):
            if header and any(kw in str(header).lower().replace(' ', '') for kw in keywords):
                if field not in mapping:
                    mapping[field] = i
                break
    return mapping


@admin_import_bp.route("/telecharger_modele")
@admin_required
def telecharger_modele_excel():
    etablissement_id = session["etablissement_id"]
    armoires = db.session.execute(db.select(Armoire.nom).filter_by(etablissement_id=etablissement_id).order_by(Armoire.nom)).scalars().all()
    categories = db.session.execute(db.select(Categorie.nom).filter_by(etablissement_id=etablissement_id).order_by(Categorie.nom)).scalars().all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventaire"
    ws.append(["Nom", "Type", "CMR", "Quantite", "Seuil", "Armoire", "Categorie", "Date Peremption", "Image URL"])
    ws_data = wb.create_sheet("Data_Listes")
    ws_data.sheet_state = "hidden"
    for i, nom in enumerate(armoires, 1): ws_data.cell(row=i, column=1, value=nom)
    for i, nom in enumerate(categories, 1): ws_data.cell(row=i, column=2, value=nom)
    dv_type = DataValidation(type="list", formula1='"materiel,produit"', allow_blank=True)
    ws.add_data_validation(dv_type)
    dv_type.add("B2:B1000")
    dv_cmr = DataValidation(type="list", formula1='"oui,non"', allow_blank=True)
    ws.add_data_validation(dv_cmr)
    dv_cmr.add("C2:C1000")
    if armoires:
        ws.add_data_validation(dv)
        dv.add("F2:F1000")
    if categories:
        ws.add_data_validation(dv)
        dv.add("G2:G1000")
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="modele_import.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@admin_import_bp.route("/importer", methods=['GET'])
@admin_required
def importer_page():
    return render_template("admin_import.html", breadcrumbs=[
        {'text': 'Tableau de Bord', 'url': url_for('inventaire.index')},
        {'text': 'Administration', 'url': url_for('admin.admin')},
        {'text': 'Importation en masse', 'url': None}
    ])


@admin_import_bp.route("/analyser_excel", methods=['POST'])
@admin_required
@limiter.limit("20 per minute")
def analyser_excel():
    """Etape 1+2 : Analyse le fichier et retourne colonnes + apercu + mapping suggere"""
    if 'fichier_excel' not in request.files:
        return jsonify({'success': False, 'error': 'Aucun fichier fourni'}), 400
    fichier = request.files['fichier_excel']
    if not fichier.filename.endswith('.xlsx'):
        return jsonify({'success': False, 'error': 'Format invalide — .xlsx uniquement'}), 400
    fichier.seek(0, 2)
    if fichier.tell() > MAX_FILE_SIZE:
        return jsonify({'success': False, 'error': 'Fichier trop volumineux (max 5 Mo)'}), 400
    fichier.seek(0)
    try:
        wb = load_workbook(fichier, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        if not any(headers):
            return jsonify({'success': False, 'error': 'La première ligne doit contenir les en-têtes de colonnes'}), 400
        apercu = []
        for row in ws.iter_rows(min_row=2, max_row=6, values_only=True):
            if any(cell is not None for cell in row):
                apercu.append([str(v) if v is not None else '' for v in row])
        mapping_suggere = detect_mapping(headers)
        etablissement_id = session['etablissement_id']
        armoires = db.session.execute(db.select(Armoire.nom).filter_by(etablissement_id=etablissement_id)).scalars().all()
        categories = db.session.execute(db.select(Categorie.nom).filter_by(etablissement_id=etablissement_id)).scalars().all()
        return jsonify({
            'success': True,
            'headers': [str(h) if h else '' for h in headers],
            'apercu': apercu,
            'mapping_suggere': mapping_suggere,
            'nb_lignes': ws.max_row - 1,
            'armoires': list(armoires),
            'categories': list(categories)
        })
    except Exception as e:
        current_app.logger.error(f'Erreur analyse Excel: {e}', exc_info=True)
        return jsonify({'success': False, 'error': 'Erreur lors de la lecture du fichier'}), 500


@admin_import_bp.route("/confirmer_import", methods=['POST'])
@admin_required
@limiter.limit("5 per minute")
def confirmer_import():
    """Etape 3 : Recoit le mapping + fichier, previsualise ou confirme l'import"""
    etablissement_id = session['etablissement_id']
    if 'fichier_excel' not in request.files:
        return jsonify({'success': False, 'error': 'Fichier manquant'}), 400
    fichier = request.files['fichier_excel']
    mode = request.form.get('mode', 'preview')
    creer_manquants = request.form.get('creer_manquants', 'false') == 'true'
    try:
        mapping = json.loads(request.form.get('mapping', '{}'))
    except Exception:
        return jsonify({'success': False, 'error': 'Mapping invalide'}), 400
    if 'nom' not in mapping:
        return jsonify({'success': False, 'error': 'La colonne Nom est obligatoire'}), 400
    try:
        wb = load_workbook(fichier, data_only=True)
        ws = wb.active
        armoires_map = {a.nom.lower().strip(): a for a in db.session.execute(db.select(Armoire).filter_by(etablissement_id=etablissement_id)).scalars().all()}
        categories_map = {c.nom.lower().strip(): c for c in db.session.execute(db.select(Categorie).filter_by(etablissement_id=etablissement_id)).scalars().all()}
        existing_noms = {o.nom.lower() for o in db.session.execute(db.select(Objet).filter_by(etablissement_id=etablissement_id)).scalars().all()}
        resultats = []
        objets_a_creer = []
        armoires_a_creer = set()
        categories_a_creer = set()

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if i > MAX_IMPORT_ROWS + 1:
                break
            if not any(row):
                continue

            def get_val(field):
                idx = mapping.get(field)
                if idx is None:
                    return None
                try:
                    return row[int(idx)]
                except Exception:
                    return None

            nom = get_val('nom')
            if not nom or str(nom).strip() == '':
                resultats.append({'ligne': i, 'statut': 'erreur', 'message': 'Nom manquant', 'nom': '—'})
                continue
            nom = str(nom).strip()
            if nom.lower() in existing_noms:
                resultats.append({'ligne': i, 'statut': 'ignore', 'message': 'Déjà existant', 'nom': nom})
                continue

            qte_raw = get_val('quantite')
            seuil_raw = get_val('seuil')
            arm_nom = str(get_val('armoire') or '').strip()
            cat_nom = str(get_val('categorie') or '').strip()
            type_objet = str(get_val('type_objet') or 'materiel').lower().strip()
            if type_objet not in ['materiel', 'produit']:
                type_objet = 'materiel'
            is_cmr_raw = str(get_val('is_cmr') or 'non').lower().strip()
            is_cmr = is_cmr_raw in ['oui', 'yes', '1', 'true', 'x']
            unite_raw = str(get_val('unite') or '').strip()
            unite = unite_raw if unite_raw in ['mL', 'L', 'g', 'kg', 'unité'] else ('unité' if type_objet == 'materiel' else 'mL')
            date_val = get_val('date_peremption')
            date_perim = None
            if date_val:
                from datetime import datetime as dt
                if hasattr(date_val, 'date'):
                    date_perim = date_val.date()
                else:
                    try:
                        date_perim = dt.strptime(str(date_val).split(' ')[0], '%Y-%m-%d').date()
                    except Exception:
                        pass

            erreurs = []
            if qte_raw is None:
                erreurs.append('Quantité manquante')
            if not arm_nom:
                erreurs.append('Armoire manquante')
            elif arm_nom.lower() not in armoires_map:
                if creer_manquants:
                    armoires_a_creer.add(arm_nom)
                else:
                    erreurs.append(f'Armoire inconnue : {arm_nom}')
            if not cat_nom:
                erreurs.append('Catégorie manquante')
            elif cat_nom.lower() not in categories_map:
                if creer_manquants:
                    categories_a_creer.add(cat_nom)
                else:
                    erreurs.append(f'Catégorie inconnue : {cat_nom}')

            if erreurs:
                resultats.append({'ligne': i, 'statut': 'erreur', 'message': ', '.join(erreurs), 'nom': nom})
            else:
                resultats.append({'ligne': i, 'statut': 'ok', 'message': 'Prêt à importer', 'nom': nom,
                                   'type_objet': type_objet, 'is_cmr': is_cmr})
                objets_a_creer.append({
                    'nom': nom, 'quantite': qte_raw, 'seuil': seuil_raw or 1,
                    'arm_nom': arm_nom, 'cat_nom': cat_nom, 'type_objet': type_objet,
                    'is_cmr': is_cmr, 'unite': unite, 'date_perim': str(date_perim) if date_perim else None
                })

        if mode == 'preview':
            return jsonify({
                'success': True,
                'resultats': resultats,
                'nb_ok': sum(1 for r in resultats if r['statut'] == 'ok'),
                'nb_erreurs': sum(1 for r in resultats if r['statut'] == 'erreur'),
                'nb_ignores': sum(1 for r in resultats if r['statut'] == 'ignore'),
                'armoires_a_creer': list(armoires_a_creer),
                'categories_a_creer': list(categories_a_creer)
            })

        # MODE IMPORT REEL
        created = 0
        for arm_nom in armoires_a_creer:
            a = Armoire(nom=arm_nom, etablissement_id=etablissement_id)
            db.session.add(a)
            db.session.flush()
            armoires_map[arm_nom.lower()] = a
        for cat_nom in categories_a_creer:
            c = Categorie(nom=cat_nom, etablissement_id=etablissement_id)
            db.session.add(c)
            db.session.flush()
            categories_map[cat_nom.lower()] = c
        for obj in objets_a_creer:
            arm = armoires_map.get(obj['arm_nom'].lower())
            cat = categories_map.get(obj['cat_nom'].lower())
            if not arm or not cat:
                continue
            try:
                qte = int(float(str(obj['quantite'])))
            except Exception:
                qte = 0
            try:
                seuil = int(float(str(obj['seuil'])))
            except Exception:
                seuil = 1
            date_perim = None
            if obj['date_perim']:
                from datetime import datetime as dt
                try:
                    date_perim = dt.strptime(obj['date_perim'], '%Y-%m-%d').date()
                except Exception:
                    pass
            db.session.add(Objet(
                nom=obj['nom'], quantite_physique=qte, seuil=seuil,
                armoire_id=arm.id, categorie_id=cat.id,
                type_objet=obj['type_objet'], is_cmr=obj['is_cmr'],
                unite=obj['unite'], date_peremption=date_perim,
                etablissement_id=etablissement_id
            ))
            created += 1

        db.session.commit()
        invalidate_alertes_cache(etablissement_id)
        cache.delete(f'armoires_{etablissement_id}')
        cache.delete(f'categories_{etablissement_id}')
        log_action('import_excel', f'{created} objets importés via assistant')
        return jsonify({'success': True, 'created': created})

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Erreur import Excel: {e}', exc_info=True)
        return jsonify({'success': False, 'error': "Erreur technique lors de l'import"}), 500

# MODULE RAPPORTS & ACTIVITÉ (Version Durcie)
# ============================================================

# Constantes de sécurité pour les rapports
MAX_EXPORT_DAYS = 366      # Limite la plage à 1 an
MAX_EXPORT_ROWS = 5000     # Limite le nombre de lignes pour éviter le crash RAM (ReportLab/OpenPyXL)
ALLOWED_FORMATS = {'excel', 'pdf'}

@admin_import_bp.route("/rapports")
@admin_required
def rapports():
    etablissement_id = session['etablissement_id']
    
    # 1. Récupération des paramètres d'URL
    page = request.args.get('page', 1, type=int)
    filtre_action = request.args.get('action', 'all') # all, creation, modification, deplacement, suppression
    search_query = request.args.get('q', '').strip()
    
    # 2. Construction de la requête de base
    stmt = (
        db.select(Historique, Utilisateur.nom_utilisateur, Objet.nom.label('objet_nom'))
        .outerjoin(Utilisateur, Historique.utilisateur_id == Utilisateur.id)
        .outerjoin(Objet, Historique.objet_id == Objet.id)
        .filter(Historique.etablissement_id == etablissement_id)
        .order_by(Historique.timestamp.desc())
    )

    # 3. Application des filtres
    if filtre_action == 'creation':
        stmt = stmt.filter(Historique.action == 'Création')
    elif filtre_action == 'suppression':
        stmt = stmt.filter(Historique.action == 'Suppression')
    elif filtre_action == 'modification':
        # Exclure les déplacements (qui sont techniquement des modifs)
        stmt = stmt.filter(
            Historique.action == 'Modification',
            ~Historique.details.ilike('%Déplacé%'),
            ~Historique.details.ilike('%Armoire%')
        )
    elif filtre_action == 'deplacement':
        stmt = stmt.filter(
            Historique.action == 'Modification',
            (Historique.details.ilike('%Déplacé%') | Historique.details.ilike('%Armoire%'))
        )

    # 4. Recherche textuelle
    if search_query:
        stmt = stmt.filter(
            (Historique.details.ilike(f'%{search_query}%')) |
            (Objet.nom.ilike(f'%{search_query}%')) |
            (Utilisateur.nom_utilisateur.ilike(f'%{search_query}%'))
        )

    # 5. Pagination (50 items par page)
    pagination = db.paginate(stmt, page=page, per_page=50)

    # 6. Formatage des données pour la vue
    logs = []
    for item in pagination.items:
        # Détection du type pour gérer le cas où paginate renvoie un objet ou un tuple
        if isinstance(item, Historique):
            # Cas : Modèle seul
            h = item
            user_name = h.utilisateur.nom_utilisateur if h.utilisateur else "Utilisateur supprimé"
            
            # Récupération manuelle du nom de l'objet (si nécessaire)
            obj_name = "Objet supprimé"
            if h.objet_id:
                obj = db.session.get(Objet, h.objet_id)
                if obj: obj_name = obj.nom
        else:
            # Cas : Tuple (Historique, nom_user, nom_objet)
            h, user_name, obj_name = item

        # Détection du type pour les badges (si pas déjà filtré)
        type_badge = 'secondary'
        if h.action == 'Création': type_badge = 'success'
        elif h.action == 'Suppression': type_badge = 'danger'
        elif 'Déplacé' in (h.details or ''): type_badge = 'info'
        elif h.action == 'Modification': type_badge = 'warning'

        logs.append({
            'date': h.timestamp,
            'user': user_name or "Utilisateur supprimé",
            'objet': obj_name or "Objet supprimé",
            'action': h.action,
            'details': h.details,
            'type_badge': type_badge
        })

    breadcrumbs=[{'text': 'Tableau de Bord', 'url': url_for('inventaire.index')}, {'text': 'Administration', 'url': url_for('admin.admin')}, {'text': 'Rapport d\'activité', 'url': None}]

    return render_template("rapports.html", 
                           logs=logs, 
                           pagination=pagination, 
                           filtre_actif=filtre_action,
                           search_query=search_query,
                           breadcrumbs=breadcrumbs)


@admin_import_bp.route("/exporter_rapports", methods=['GET'])
@admin_required
@limiter.limit("5 per minute")
def exporter_rapports():
    etablissement_id = session['etablissement_id']
    
    # 1. Récupération des paramètres
    date_debut_str = request.args.get('date_debut')
    date_fin_str = request.args.get('date_fin')
    format_type = request.args.get('format')
    group_by = request.args.get('group_by', 'date')
    
    # NOUVEAU : Récupération des actions cochées (liste)
    # Flask récupère les checkbox multiples avec getlist
    selected_actions = request.args.getlist('actions') 

    if not all([date_debut_str, date_fin_str, format_type]):
        flash("Paramètres manquants.", "warning")
        return redirect(url_for('admin_import.rapports'))

    if format_type not in ALLOWED_FORMATS:
        flash("Format non supporté.", "error")
        return redirect(url_for('admin_import.rapports'))

    try:
        # 2. Parsing Dates
        date_debut = datetime.strptime(date_debut_str, '%Y-%m-%d')
        date_fin = datetime.strptime(date_fin_str, '%Y-%m-%d')
        date_fin = date_fin.replace(hour=23, minute=59, second=59)

        if date_debut > date_fin:
            flash("Dates incohérentes.", "warning")
            return redirect(url_for('admin_import.rapports'))
            
        if (date_fin - date_debut).days > MAX_EXPORT_DAYS:
            flash(f"Période limitée à {MAX_EXPORT_DAYS} jours.", "warning")
            return redirect(url_for('admin_import.rapports'))

        # 3. Construction Requête
        query = db.select(
            Historique,
            Utilisateur.nom_utilisateur,
            Objet.nom.label('objet_nom')
        ).outerjoin(Utilisateur, Historique.utilisateur_id == Utilisateur.id)\
         .outerjoin(Objet, Historique.objet_id == Objet.id)\
         .filter(Historique.etablissement_id == etablissement_id)\
         .filter(Historique.timestamp >= date_debut)\
         .filter(Historique.timestamp <= date_fin)

        # --- LOGIQUE DE FILTRE ET TRI ---
        if group_by == 'action':
            # Si des actions spécifiques sont cochées, on filtre
            if selected_actions:
                query = query.filter(Historique.action.in_(selected_actions))
            
            # On trie par Action puis par Date
            query = query.order_by(Historique.action.asc(), Historique.timestamp.desc())
        else:
            # Tri chronologique standard
            query = query.order_by(Historique.timestamp.desc())

        # Protection RAM
        query = query.limit(MAX_EXPORT_ROWS)
        resultats = db.session.execute(query).all()

        if not resultats:
            flash("Aucune donnée trouvée pour ces critères.", "info")
            return redirect(url_for('admin_import.rapports'))

        # 4. Préparation Données
        data_export = []
        for h, user_name, obj_name in resultats:
            data_export.append({
                'date': h.timestamp.strftime('%d/%m/%Y'),
                'heure': h.timestamp.strftime('%H:%M'),
                'utilisateur': user_name or "Inconnu",
                'action': h.action,
                'objet': obj_name or "-",
                'details': h.details or ""
            })

        # Métadonnées enrichies
        filtre_info = "Tous types"
        if group_by == 'action' and selected_actions:
            filtre_info = ", ".join(selected_actions)

        params = get_etablissement_params(session.get('etablissement_id'))
        logo_url = params.get('logo_url')
        logo_path = os.path.join(current_app.root_path, 'static', logo_url.lstrip('/static/')) if logo_url else None
        metadata = {
            'etablissement': session.get('nom_etablissement', 'LabFlow'),
            'periode': f"Du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}",
            'total': len(data_export),
            'date_generation': datetime.now().strftime('%d/%m/%Y à %H:%M'),
            'filtre': filtre_info,
            'logo_path': logo_path
        }

        log_action('export_rapport', f"Format: {format_type}, Rows: {len(data_export)}")

        if format_type == 'excel':
            return generer_rapport_excel(data_export, metadata)
        else:
            return generer_rapport_pdf(data_export, metadata)

    except Exception as e:
        current_app.logger.error(f"Erreur export: {e}", exc_info=True)
        flash("Erreur technique lors de la génération.", "error")
        return redirect(url_for('admin_import.rapports'))



